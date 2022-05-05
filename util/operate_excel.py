# -*- coding = utf-8 -*-
import os

import xlrd
from openpyxl import Workbook

from config import constants
from util.logger import Logger as logs
import pandas as pd


class OperateExcel(object):

    def __init__(self):
        self.log = logs
        pass

    # 写入excel信息
    def write_excel_info(self, write_info_list, file_path):
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Sheet1"
        try:
            write_rows = 1
            for write_dict in write_info_list:
                write_cols = 1
                if write_rows == 1:
                    for write_table in write_dict.keys():
                        sheet.cell(write_rows, write_cols).value = write_table
                        write_cols += 1
                write_cols = 1
                write_rows += 1
                for write_info in write_dict.values():
                    sheet.cell(write_rows, write_cols).value = write_info
                    write_cols += 1
            wb.save(file_path)
            wb.close()
            return True
        except Exception as error:
            # self.log.error("创建excel文件错误{}".format(error))
            return False

    # 获取excel信息
    def get_excel_info(self, file_name):
        data = xlrd.open_workbook(file_contents=file_name)
        table = data.sheets()[0]
        try:
            # 获取总行数、总列数
            nrows = table.nrows
            if nrows > 1:
                # 获取第一列的内容，列表格式
                keys = table.row_values(0)
                data_list = []
                # 获取每一行的内容，列表格式
                for col in range(1, nrows):
                    values = table.row_values(col)
                    # keys，values这两个列表一一对应来组合转换为字典
                    data_dict = dict(zip(keys, values))
                    data_list.append(data_dict)
                return data_list
            else:
                self.log.error("表格没有读取到数据")
                return None
        except Exception as error:
            self.log.error("读取excel信息报错{}".format(error))
            return None

    # 获取excel信息
    def get_recording_list(self, file_data, file_name, index_x, index_y, replace_num, file_recording_set):
        if file_name is None or file_name == '':
            return
        try:
            file = xlrd.open_workbook(file_contents=file_data)
            sheet = file.sheet_by_name(file_name)
            # 获取总行数、总列数
            nrows = sheet.nrows
            for i in range(index_x, nrows):
                number = sheet.cell(i, index_y).value
                if number is None or number == '':
                    continue
                if replace_num != '' and replace_num != constants.RECORD_ROUND_REP:
                    number = str(number).replace(constants.RECORD_ROUND_REP, replace_num)
                file_recording_set.add(number)
        except Exception as ex:
            self.log.error(ex)

    # 获取excel信息
    def get_recording_list_pandas(self, file_data, file_name, replace_num, file_recording_set):
        if file_name is None or file_name == '':
            return
        try:
            self.log.info("获取文件【{}】录音编号".format(file_name))
            df = pd.read_excel(file_data, sheet_name=file_name, header=0)
            df_list = df['录音编号'].unique()
            for number in df_list:
                if pd.isna(number):
                    continue
                if replace_num != '' and replace_num != constants.RECORD_ROUND_REP:
                    number = str(number).replace(constants.RECORD_ROUND_REP, replace_num)
                file_recording_set.add(number)
        except Exception as ex:
            self.log.error(ex)


    @staticmethod
    def define_excel_format(info_type, faq_title, faq_extend, faq_answer, info_list):
        faq_dict = {
            "所属类型": info_type,
            "标准问": faq_title,
            "扩展问": faq_extend,
            "答案": faq_answer,
        }
        info_list.append(faq_dict)

    def createFile(self, file_path):
        if os.path.exists(file_path) is False:
            os.makedirs(file_path)
            return file_path
        else:
            return file_path
