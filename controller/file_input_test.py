import re
import time

import openpyxl
import xlrd
from openpyxl.utils import range_boundaries

from config import constants
from util.merge_cell_util import get_cell_type


def read_file1():
    readBook = xlrd.open_workbook(r'../excel/京东养老金开户bot.xlsx')
    # sheetFaq = readBook.sheet_by_index(0)
    sheetFaq = readBook.sheet_by_name("京东养老金带操作D版-0719(FZ0712)")
    start_i = 0
    end = 120
    for i in range(start_i, end):
        merged = sheetFaq.merged_cells
        label2 = get_cell_type(i, constants.FLOW_NAME, merged, sheetFaq)
        print(label2)

def read_file3(path, sheetname):
    start_time = time.time()
    # 打开Excel文件
    workbook = openpyxl.load_workbook('../excel/{}'.format(path))
    print("耗时：{}".format(time.time() - start_time))
    # 获取第一个工作表
    worksheet = workbook[sheetname]
    listw = []
    # 遍历第二列单元格
    for cell in worksheet['B']:
        temp = []
        if cell.value:
            # 如果单元格是合并单元格，打印合并单元格的值
            if cell.coordinate in worksheet.merged_cells:
                for merged_cell in worksheet.merged_cells:
                    if cell.coordinate in merged_cell:
                        min_row = merged_cell.min_row
                        max_row = merged_cell.max_row
                        name = worksheet[min_row][merged_cell.min_col-1].value
                        # print("合并{} row_min:{} row_max:{}".format(name, min_row, max_row))
                        temp = [name, min_row, max_row]
            else:
                temp = [cell.value, cell.row, cell.row]
        if len(temp) > 0 and "DM使用" not in temp[0]:
            listw.append(temp)
    for res in listw:
        print("{}".format(res))
    return listw


def read_file_xrld(path, sheetname):
    start_time = time.time()
    # 打开Excel文件
    workbook = xlrd.open_workbook('../excel/{}'.format(path))
    print("耗时：{}".format(time.time() - start_time))
    # 获取第一个工作表
    worksheet = workbook.sheet_by_name(sheetname)
    all_B_cells = []
    col = 1
    merged_cells = worksheet.merged_cells
    for row in range(worksheet.nrows):
        temp = None
        # 获取B列单元格的内容
        cell_value = worksheet.cell_value(row, 1)
        # 判断该单元格是否是合并单元格
        for merged_cell in merged_cells:
            if row == merged_cell[0] and col == merged_cell[2]:
                # 如果该单元格是合并单元格，则获取合并单元格的内容
                merged_cell_value = worksheet.cell_value(merged_cell[0], merged_cell[2])
                temp = [merged_cell_value, merged_cell[0]+1, merged_cell[1]]
                break
        else:
            # 如果该单元格不是合并单元格，则直接打印单元格的内容
            if cell_value:
                temp = [cell_value, row+1, row+1]
        if temp and "DM使用" not in temp[0]:
            all_B_cells.append(temp)

    for res in all_B_cells:
        print("{}".format(res[0]))
    return all_B_cells


def read_file():
    # 导入openpyxl库
    import openpyxl

    # 打开Excel文件
    wb = openpyxl.load_workbook('../excel/京东养老金开户bot.xlsx')
    # # 选择工作表
    # sheet = wb.active
    sheet = wb['京东养老金带操作D版-0719(FZ0712)']
    cell_value = sheet.cell(row=3, column=1).value
    # 遍历所有合并单元格
    for merged_cell in sheet.merged_cells:
        # 获取合并单元格的坐标
        coord = merged_cell.coord
        # 将坐标转换为行列号
        start_row, start_col, end_row, end_col = range_boundaries(coord)
        # 获取合并单元格的内容
        cell_value = sheet.cell(row=1, column=start_col).value
        # 输出结果
        print(f"合并单元格{start_row}:{start_col}-{end_row}:{end_col}的内容为：{cell_value}")


def test():
    # reverseOrder = "1.x.y.4\n(x=1,6)(y=1,6)"
    reverseOrder = str(1)

    if "." not in reverseOrder:
        return ""
    # match_num = re.search("(.*?)\.\d", reverseOrder)
    # if match_num is None:
    #     return None
    # order = match_num.group()
    last_dot_index = str(reverseOrder).rfind(".")
    order = reverseOrder[:last_dot_index]

    print(order)

    letters = re.findall(r'\(([a-zA-Z]=\d).*?\)', reverseOrder)
    if letters is not None and len(letters) > 0:
        letter = letters[len(letters) - 1]
        if "=" in letter:
            kw = letter.split("=")
            order = order.replace(kw[0], kw[1])
    print(order)

def test2():
    reverseOrder = "1.x.y.4\n(x=1,6)(y=1,6)"
    # reverseOrder = "1.x.y.4"
    match_num = re.search("(.*?\.\d)", reverseOrder)
    if match_num:
        order = match_num.group(1)
    else:
        order = reverseOrder
    print(order)


if __name__ == '__main__':
    # path = "金条-首贷-已激活-纯机-新版.xlsx"
    # sheetName = "金条-首贷-已激活-纯机-新版-0822"
    # # read_file3(path, sheetName)
    # read_file_xrld(path, sheetName)

    test()
    pass
